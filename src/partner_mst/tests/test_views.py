import csv
import io
import json
from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from django.urls import reverse
from partner_mst.models import Partner
from django.core.management import call_command
from bs4 import BeautifulSoup
from django.utils import timezone
from tenant_mst.models import Tenant
from sales_order.models import SalesOrder
from django.contrib import messages
from partner_mst.views import DATA_COLUMNS
from openpyxl import load_workbook


class PartnerViewTests(TestCase):
    def setUp(self):
        '''共通データ作成'''
        # テストクライアント生成
        self.client = Client()
        
        # テストデータ投入
        call_command('loaddata', 'test_tenants.json')
        call_command('loaddata', 'test_registers.json')
        call_command('loaddata', 'test_partners.json')
        
        # システムユーザーで実施
        self.user = get_user_model().objects.get(pk=1)
        self.client.login(email='system@example.com', password='pass')

    # -------------------
    # 初期表示
    # -------------------
    def test_1_1_1(self):
        '''初期表示（データあり）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'))
        
        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # レスポンス内容確認
        self.assertContains(response, '株式会社アルファ')
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 8)  # 件数
        self.assertTrue(all(tid == 1 for tid in list.values_list('tenant_id', flat=True)))  # テナントID
        
    def test_1_1_2(self):
        '''初期表示（データなし）'''
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'))

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 0)  # 件数
        
    # -------------------
    # 検索処理
    # -------------------
    def test_2_1_1(self):
        '''検索処理（取引先名称）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_partner_name': 'アルファ'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('株式会社アルファ', list[0].partner_name)  # データ値
        self.assertEqual('アルファ', response.context['search_partner_name'])  # 検索フォームの入力値
        
    def test_2_1_2(self):
        '''検索処理（担当者名）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_contact_name': '佐藤'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('佐藤 次郎', list[0].contact_name)  # データ値
        self.assertEqual('佐藤', response.context['search_contact_name'])  # 検索フォームの入力値

    def test_2_1_3(self):
        '''検索処理（電話番号）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_tel_number': '0356781234'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('0356781234', list[0].tel_number)  # データ値
        self.assertEqual('0356781234', response.context['search_tel_number'])  # 検索フォームの入力値
        
    def test_2_1_4(self):
        '''検索処理（メールアドレス）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_email': 'cs.'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('client@cs.co.jp', list[0].email)  # データ値
        self.assertEqual('cs.', response.context['search_email'])  # 検索フォームの入力値
        
    def test_2_1_5(self):
        '''検索処理（都道府県）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_address': '北海道'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('北海道', list[0].state)  # データ値
        self.assertEqual('北海道', response.context['search_address'])  # 検索フォームの入力値
        
    def test_2_1_6(self):
        '''検索処理（市町村区）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_address': '千代田'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('千代田区千代田1-1', list[0].city)  # データ値
        self.assertEqual('千代田', response.context['search_address'])  # 検索フォームの入力値
        
    def test_2_1_7(self):
        '''検索処理（住所）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_address': 'フーズビル'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('フーズビル', list[0].address)  # データ値
        self.assertEqual('フーズビル', response.context['search_address'])  # 検索フォームの入力値
        
    def test_2_1_8(self):
        '''検索処理（住所2）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_address': 'セカンドアドレス'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('セカンドアドレス', list[0].address2)  # データ値
        self.assertEqual('セカンドアドレス', response.context['search_address'])  # 検索フォームの入力値

    def test_2_1_9(self):
        '''検索処理（非所属テナント）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_address': 'グローバル貿易株式会社'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 0)  # 件数
        self.assertEqual('グローバル貿易株式会社', response.context['search_address'])  # 検索フォームの入力値
        
    def test_2_1_10(self):
        '''検索処理（取引先区分）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:list'), {'search_partner_type': 'both'})

        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # 取得データ確認
        list = response.context['partners']
        self.assertEqual(list.count(), 1)  # 件数
        self.assertEqual('both', response.context['search_partner_type'])  # 検索フォームの入力値
        
    # -------------------
    # 登録処理
    # -------------------
    def test_3_1_1(self):
        '''登録画面表示（正常系）'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:create'))
        
        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # JSONをロード
        data = json.loads(response.content)
        html = data['html']

        # BeautifulSoupでHTMLを解析
        soup = BeautifulSoup(html, 'html.parser')
        modal_title = soup.select_one('.modal-title').get_text(strip=True)

        # モーダルタイトル確認
        self.assertEqual(modal_title, '取引先: 新規登録')
        
    def test_3_2_1(self):
        '''
        登録処理（正常系：全項目に設定）
        '''
        url = reverse('partner_mst:create')
        data = {
            'partner_name': 'テスト株式会社',
            'partner_name_kana': 'テストカブシキガイシャ',
            'partner_type': 'customer',
            'contact_name': '山田太郎',
            'tel_number': '0312345678',
            'email': 'test@example.com',
            'postal_code': '1000001',
            'state': '東京都',
            'city': '千代田区',
            'address': '丸の内1-1-1',
            'address2': 'ビル3F',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB登録確認
        partner = Partner.objects.get(partner_name='テスト株式会社')
        self.assertEqual(partner.partner_name, 'テスト株式会社')
        self.assertEqual(partner.partner_name_kana, 'テストカブシキガイシャ')
        self.assertEqual(partner.partner_type, 'customer')
        self.assertEqual(partner.contact_name, '山田太郎')
        self.assertEqual(partner.tel_number, '0312345678')
        self.assertEqual(partner.email, 'test@example.com')
        self.assertEqual(partner.postal_code, '1000001')
        self.assertEqual(partner.state, '東京都')
        self.assertEqual(partner.city, '千代田区')
        self.assertEqual(partner.address, '丸の内1-1-1')
        self.assertEqual(partner.address2, 'ビル3F')
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, self.user)
        self.assertEqual(partner.update_user, self.user)
        self.assertEqual(partner.tenant, self.user.tenant)
        # created_at / updated_at の存在確認と現在時刻との差を確認（±5秒以内ならOK）
        now = timezone.now()
        self.assertLessEqual(abs((partner.created_at - now).total_seconds()), 5)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_3_2_2(self):
        '''登録処理（正常系：必須項目のみ）'''
        url = reverse('partner_mst:create')
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB登録確認
        partner = Partner.objects.get(partner_name='テスト株式会社')
        self.assertEqual(partner.partner_name, 'テスト株式会社')
        self.assertEqual(partner.partner_type, 'customer')
        self.assertEqual(partner.email, 'test@example.com')
        self.assertEqual(partner.create_user, self.user)
        self.assertEqual(partner.update_user, self.user)
        self.assertEqual(partner.tenant, self.user.tenant)
        # created_at / updated_at の存在確認と現在時刻との差を確認（±5秒以内ならOK）
        now = timezone.now()
        self.assertLessEqual(abs((partner.created_at - now).total_seconds()), 5)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_3_2_3(self):
        '''登録処理（正常系：最大文字数）'''
        url = reverse('partner_mst:create')
        partner_name = 'a' * 255
        partner_name_kana = 'b' * 255
        partner_type = 'customer'
        contact_name = 'c' * 255
        tel_number = '1' * 15
        email = 'test@example.com'
        postal_code = '1' * 7
        state = 'f' * 5
        city = 'g' * 255
        address = 'h' * 255
        address2 = 'i' * 255
        
        data = {
            'partner_name': partner_name,
            'partner_name_kana': partner_name_kana,
            'partner_type': partner_type,
            'contact_name': contact_name,
            'tel_number': tel_number,
            'email': email,
            'postal_code': postal_code,
            'state': state,
            'city': city,
            'address': address,
            'address2': address2,
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB登録確認
        partner = Partner.objects.get(partner_name=partner_name)
        self.assertEqual(partner.partner_name, partner_name)
        self.assertEqual(partner.partner_name_kana, partner_name_kana)
        self.assertEqual(partner.partner_type, partner_type)
        self.assertEqual(partner.contact_name, contact_name)
        self.assertEqual(partner.tel_number, tel_number)
        self.assertEqual(partner.email, email)
        self.assertEqual(partner.postal_code, postal_code)
        self.assertEqual(partner.state, state)
        self.assertEqual(partner.city, city)
        self.assertEqual(partner.address, address)
        self.assertEqual(partner.address2, address2)
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, self.user)
        self.assertEqual(partner.update_user, self.user)
        self.assertEqual(partner.tenant, self.user.tenant)
        # created_at / updated_at の存在確認と現在時刻との差を確認（±5秒以内ならOK）
        now = timezone.now()
        self.assertLessEqual(abs((partner.created_at - now).total_seconds()), 5)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_3_2_4(self):
        '''登録処理（正常系：異なるテナントで同じ取引先名称）'''
        # テナントを2つ作成
        self.tenant1 = Tenant.objects.create(
            tenant_name='テナントA',
            representative_name='代表A',
            contact_email='a@example.com'
        )
        self.tenant2 = Tenant.objects.create(
            tenant_name='テナントB',
            representative_name='代表B',
            contact_email='b@example.com'
        )

        # ユーザーを作成してテナント1に所属
        self.user1 = get_user_model().objects.create_user(
            username='user1',
            email='user1@example.com',
            password='pass',
            privilege=0,
            tenant=self.tenant1
        )

        # ユーザーを作成してテナント2に所属
        self.user2 = get_user_model().objects.create_user(
            username='user2',
            email='user2@example.com',
            password='pass',
            privilege=0,
            tenant=self.tenant2
        )
        
        url = reverse('partner_mst:create')
        data = {
            'partner_name': '重複テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com'
        }

        # テナント1で登録
        self.client.login(email='user1@example.com', password='pass')
        response = self.client.post(
            url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # テナント2で同じ名前を登録
        self.client.logout()
        self.client.login(email='user2@example.com', password='pass')
        response = self.client.post(
            url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB確認
        partners_t1 = Partner.objects.filter(tenant=self.tenant1, partner_name='重複テスト株式会社')
        partners_t2 = Partner.objects.filter(tenant=self.tenant2, partner_name='重複テスト株式会社')

        # 件数確認
        self.assertEqual(partners_t1.count(), 1)
        self.assertEqual(partners_t2.count(), 1)

    def test_3_3_1(self):
        '''
        登録処理（異常系：取引先名称必須チェック）
        '''
        url = reverse('partner_mst:create')
        data = {
            'partner_name_kana': 'テストカブシキガイシャ',
            'partner_type': 'customer',
            'contact_name': '山田太郎',
            'tel_number': '0312345678',
            'email': 'test@example.com',
            'postal_code': '1000001',
            'state': '東京都',
            'city': '千代田区',
            'address': '丸の内1-1-1',
            'address2': 'ビル3F',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])
        self.assertEqual(Partner.objects.filter(email='test@example.com').count(), 0)  # 件数確認
        
        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('この項目は必須です。', soup.select_one('#id_partner_name + .invalid-feedback').get_text())
        
    def test_3_3_2(self):
        '''登録処理（異常系：最大文字数超過）'''
        url = reverse('partner_mst:create')
        partner_name = 'a' * 256
        partner_name_kana = 'b' * 256
        partner_type = 'customer'
        contact_name = 'c' * 256
        tel_number = 'd' * 16
        email = 'test@example.com'
        postal_code = 'e' * 8
        state = 'f' * 6
        city = 'g' * 256
        address = 'h' * 256
        address2 = 'i' * 256
        
        data = {
            'partner_name': partner_name,
            'partner_name_kana': partner_name_kana,
            'partner_type': partner_type,
            'contact_name': contact_name,
            'tel_number': tel_number,
            'email': email,
            'postal_code': postal_code,
            'state': state,
            'city': city,
            'address': address,
            'address2': address2,
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])
        self.assertEqual(Partner.objects.filter(partner_name=partner_name).count(), 0)  # 件数確認
        
        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        
        # フィールドごとのエラーメッセージを確認
        self.assertIn('255 文字以下', soup.select_one('#id_partner_name + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_partner_name_kana + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_contact_name + .invalid-feedback').get_text())
        self.assertIn('15 文字以下', soup.select_one('#id_tel_number + .invalid-feedback').get_text())
        self.assertIn('7 文字以下', soup.select_one('#id_postal_code + .invalid-feedback').get_text())
        self.assertIn('5 文字以下', soup.select_one('#id_state + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_city + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_address + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_address2 + .invalid-feedback').get_text())


    def test_3_3_3(self):
        '''登録処理（異常系：メールアドレス形式不正）'''
        url = reverse('partner_mst:create')
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@@test',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # DBに登録されていないことを確認
        self.assertEqual(Partner.objects.filter(partner_name='テスト株式会社').count(), 0)

        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertIn('有効なメールアドレスを入力してください。', soup.select_one('#id_email + .invalid-feedback').get_text())
        

    def test_3_3_4(self):
        '''登録処理（異常系：メールアドレス必須チェック）'''
        url = reverse('partner_mst:create')
        data = {
            'partner_name_kana': 'テスト株式会社',
            'partner_type': 'customer',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])
        self.assertEqual(Partner.objects.filter(email='test@example.com').count(), 0)  # 件数確認
        
        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('この項目は必須です。', soup.select_one('#id_email + .invalid-feedback').get_text())

    def test_3_3_5(self):
        '''登録処理（異常系：取引先区分不正）'''
        url = reverse('partner_mst:create')
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'invalid',
            'email': 'test@example.com',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertEqual(Partner.objects.filter(partner_name='テスト株式会社').count(), 0)  # 件数確認
        
        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('正しく選択してください。 invalid は候補にありません。', soup.select_one('#id_partner_type + .invalid-feedback').get_text())

    def test_3_3_6(self):
        '''登録処理（異常系：同一テナント内取引先名称重複）'''
        url = reverse('partner_mst:create')

        # 先に1件目を登録
        data1 = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com',
        }
        response1 = self.client.post(
            url,
            data1,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        self.assertEqual(response1.status_code, 200)
        res_json1 = json.loads(response1.content)
        self.assertTrue(res_json1['success'])

        # 同じ取引先名称で2件目を登録
        data2 = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test2@example.com',
        }
        response2 = self.client.post(
            url,
            data2,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        self.assertEqual(response2.status_code, 200)
        res_json2 = json.loads(response2.content)
        self.assertFalse(res_json2['success'])

        # DBに2件目は登録されていないことを確認
        self.assertEqual(Partner.objects.filter(email='test@example.com').count(), 1)
        
        # エラーメッセージをHTMLから抽出
        soup = BeautifulSoup(res_json2['html'], 'html.parser')
        self.assertIn('同じ取引先名称とメールアドレスの組み合わせが既に登録されています。', soup.select_one('#id_partner_name + .invalid-feedback').get_text())
        self.assertIn('同じ取引先名称とメールアドレスの組み合わせが既に登録されています。', soup.select_one('#id_email + .invalid-feedback').get_text())

    # -------------------
    # 更新処理
    # -------------------
    def test_4_1_1(self):
        '''更新画面表示'''
        # レスポンス取得
        response = self.client.get(reverse('partner_mst:update', args=[1]))
        
        # ステータスコード確認
        self.assertEqual(response.status_code, 200)
        
        # JSONをロード
        data = json.loads(response.content)
        html = data['html']

        # BeautifulSoupでHTMLを解析
        soup = BeautifulSoup(html, 'html.parser')
        modal_title = soup.select_one('.modal-title').get_text(strip=True)

        # モーダルタイトル確認
        self.assertEqual(modal_title, '取引先更新: 株式会社アルファ')
        
    def test_4_2_1(self):
        '''
        更新処理（正常系：全項目更新）
        '''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='旧テスト株式会社',
            partner_name_kana='キュウテストカブシキガイシャ',
            partner_type='supplier',
            contact_name='佐藤一郎',
            tel_number='0311111111',
            email='old@example.com',
            postal_code='1500001',
            state='東京都',
            city='渋谷区',
            address='渋谷1-1-1',
            address2='旧ビル2F',
            create_user=self.user,
            update_user=self.user,
        )
        create_user = self.user
        
        # 事前データ作成ユーザーとは別のユーザーで実施
        self.user = get_user_model().objects.get(pk=2)
        self.client.login(email='manager@example.com', password='pass')

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_name_kana': 'テストカブシキガイシャ',
            'partner_type': 'customer',
            'contact_name': '山田太郎',
            'tel_number': '0312345678',
            'email': 'test@example.com',
            'postal_code': '1000001',
            'state': '東京都',
            'city': '千代田区',
            'address': '丸の内1-1-1',
            'address2': 'ビル3F',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB更新確認
        partner.refresh_from_db()
        self.assertEqual(partner.partner_name, 'テスト株式会社')
        self.assertEqual(partner.partner_name_kana, 'テストカブシキガイシャ')
        self.assertEqual(partner.partner_type, 'customer')
        self.assertEqual(partner.contact_name, '山田太郎')
        self.assertEqual(partner.tel_number, '0312345678')
        self.assertEqual(partner.email, 'test@example.com')
        self.assertEqual(partner.postal_code, '1000001')
        self.assertEqual(partner.state, '東京都')
        self.assertEqual(partner.city, '千代田区')
        self.assertEqual(partner.address, '丸の内1-1-1')
        self.assertEqual(partner.address2, 'ビル3F')
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, create_user)  # 作成時のユーザー
        self.assertEqual(partner.update_user, self.user)  # 更新時のユーザー
        self.assertEqual(partner.tenant, self.user.tenant)
        now = timezone.now()
        self.assertLess((now - partner.created_at).total_seconds(), 60)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_4_2_2(self):
        '''
        更新処理（正常系：全項目更新）
        '''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='旧テスト株式会社',
            partner_type='supplier',
            email='old@example.com',
            create_user=self.user,
            update_user=self.user,
        )
        create_user = self.user
        
        # 事前データ作成ユーザーとは別のユーザーで実施
        self.user = get_user_model().objects.get(pk=2)
        self.client.login(email='manager@example.com', password='pass')

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com',
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB更新確認
        partner.refresh_from_db()
        self.assertEqual(partner.partner_name, 'テスト株式会社')
        self.assertEqual(partner.partner_type, 'customer')
        self.assertEqual(partner.email, 'test@example.com')
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, create_user)  # 作成時のユーザー
        self.assertEqual(partner.update_user, self.user)  # 更新時のユーザー
        self.assertEqual(partner.tenant, self.user.tenant)
        now = timezone.now()
        self.assertLess((now - partner.created_at).total_seconds(), 60)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_4_2_3(self):
        '''
        更新処理（正常系：最大文字数）
        '''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='旧テスト株式会社',
            partner_name_kana='キュウテストカブシキガイシャ',
            partner_type='supplier',
            contact_name='佐藤一郎',
            tel_number='0311111111',
            email='old@example.com',
            postal_code='1500001',
            state='東京都',
            city='渋谷区',
            address='渋谷1-1-1',
            address2='旧ビル2F',
            create_user=self.user,
            update_user=self.user,
        )
        create_user = self.user
        
        # 事前データ作成ユーザーとは別のユーザーで実施
        self.user = get_user_model().objects.get(pk=2)
        self.client.login(email='manager@example.com', password='pass')

        url = reverse('partner_mst:update', args=[partner.id])
        
        # 更新用設定値
        partner_name = 'a' * 255
        partner_name_kana = 'b' * 255
        partner_type = 'customer'
        contact_name = 'c' * 255
        tel_number = '1' * 15
        email = 'test@example.com'
        postal_code = '1' * 7
        state = 'f' * 5
        city = 'g' * 255
        address = 'h' * 255
        address2 = 'i' * 255
        
        # 更新データ作成
        data = {
            'partner_name': partner_name,
            'partner_name_kana': partner_name_kana,
            'partner_type': partner_type,
            'contact_name': contact_name,
            'tel_number': tel_number,
            'email': email,
            'postal_code': postal_code,
            'state': state,
            'city': city,
            'address': address,
            'address2': address2,
        }

        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB登録確認
        partner = Partner.objects.get(partner_name=partner_name)
        self.assertEqual(partner.partner_name, partner_name)
        self.assertEqual(partner.partner_name_kana, partner_name_kana)
        self.assertEqual(partner.partner_type, partner_type)
        self.assertEqual(partner.contact_name, contact_name)
        self.assertEqual(partner.tel_number, tel_number)
        self.assertEqual(partner.email, email)
        self.assertEqual(partner.postal_code, postal_code)
        self.assertEqual(partner.state, state)
        self.assertEqual(partner.city, city)
        self.assertEqual(partner.address, address)
        self.assertEqual(partner.address2, address2)
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, create_user)  # 作成時のユーザー
        self.assertEqual(partner.update_user, self.user)  # 更新時のユーザー
        self.assertEqual(partner.tenant, self.user.tenant)
        now = timezone.now()
        self.assertLess((now - partner.created_at).total_seconds(), 60)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)
        
    def test_4_2_4(self):
        '''更新処理（正常系：異なるテナントで同じ取引先名称に変更）'''
        # テナントを2つ作成
        self.tenant1 = Tenant.objects.create(
            tenant_name='テナントA',
            representative_name='代表A',
            contact_email='a@example.com'
        )
        self.tenant2 = Tenant.objects.create(
            tenant_name='テナントB',
            representative_name='代表B',
            contact_email='b@example.com'
        )

        # ユーザーを作成してテナント1に所属
        self.user1 = get_user_model().objects.create_user(
            username='user1',
            email='user1@example.com',
            password='pass',
            privilege=0,
            tenant=self.tenant1
        )

        # ユーザーを作成してテナント2に所属
        self.user2 = get_user_model().objects.create_user(
            username='user2',
            email='user2@example.com',
            password='pass',
            privilege=0,
            tenant=self.tenant2
        )

        # 各テナントに取引先データを事前作成（名前は異なる）
        partner1 = Partner.objects.create(
            tenant=self.tenant1,
            partner_name='テナントA取引先',
            email='a-partner@example.com',
            partner_type='customer',
            create_user=self.user1,
            update_user=self.user1,
        )
        partner2 = Partner.objects.create(
            tenant=self.tenant2,
            partner_name='テナントB取引先',
            email='b-partner@example.com',
            partner_type='customer',
            create_user=self.user2,
            update_user=self.user2,
        )

        # 更新データ（両方とも同じ名前に更新してみる）
        data = {
            'partner_name': '重複テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com'
        }

        # テナント1で更新
        self.client.login(email='user1@example.com', password='pass')
        url1 = reverse('partner_mst:update', args=[partner1.id])
        response = self.client.post(url1, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # テナント2で更新
        self.client.logout()
        self.client.login(email='user2@example.com', password='pass')
        url2 = reverse('partner_mst:update', args=[partner2.id])
        response = self.client.post(url2, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertTrue(res_json['success'])

        # DB確認（両テナントで同じ名前が登録されている）
        partners_t1 = Partner.objects.filter(tenant=self.tenant1, partner_name='重複テスト株式会社')
        partners_t2 = Partner.objects.filter(tenant=self.tenant2, partner_name='重複テスト株式会社')

        # 件数確認
        self.assertEqual(partners_t1.count(), 1)
        self.assertEqual(partners_t2.count(), 1)

    def test_4_3_1(self):
        '''
        更新処理（異常系：取引先名称必須チェック）
        '''
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='旧テスト株式会社',
            email='old@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name_kana': 'テストカブシキガイシャ',
            'partner_type': 'customer',
            'contact_name': '山田太郎',
            'tel_number': '0312345678',
            'email': 'test@example.com',
            'postal_code': '1000001',
            'state': '東京都',
            'city': '千代田区',
            'address': '丸の内1-1-1',
            'address2': 'ビル3F',
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner.refresh_from_db()
        self.assertEqual(partner.email, 'old@example.com')

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('この項目は必須です。', soup.select_one('#id_partner_name + .invalid-feedback').get_text())

    def test_4_3_2(self):
        '''更新処理（異常系：最大文字数超過）'''
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='旧テスト株式会社',
            email='old@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner.id])
        partner_name = 'a' * 256
        partner_name_kana = 'b' * 256
        contact_name = 'c' * 256
        tel_number = 'd' * 16
        email = 'test@example.com'
        postal_code = 'e' * 8
        state = 'f' * 6
        city = 'g' * 256
        address = 'h' * 256
        address2 = 'i' * 256

        data = {
            'partner_name': partner_name,
            'partner_name_kana': partner_name_kana,
            'partner_type': 'customer',
            'contact_name': contact_name,
            'tel_number': tel_number,
            'email': email,
            'postal_code': postal_code,
            'state': state,
            'city': city,
            'address': address,
            'address2': address2,
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner.refresh_from_db()
        self.assertEqual(partner.partner_name, '旧テスト株式会社')

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertIn('255 文字以下', soup.select_one('#id_partner_name + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_partner_name_kana + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_contact_name + .invalid-feedback').get_text())
        self.assertIn('15 文字以下', soup.select_one('#id_tel_number + .invalid-feedback').get_text())
        self.assertIn('7 文字以下', soup.select_one('#id_postal_code + .invalid-feedback').get_text())
        self.assertIn('5 文字以下', soup.select_one('#id_state + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_city + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_address + .invalid-feedback').get_text())
        self.assertIn('255 文字以下', soup.select_one('#id_address2 + .invalid-feedback').get_text())

    def test_4_3_3(self):
        '''更新処理（異常系：メールアドレス形式不正）'''
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='テスト株式会社',
            email='old@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@@test',
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner.refresh_from_db()
        self.assertEqual(partner.email, 'old@example.com')

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('有効なメールアドレスを入力してください。', soup.select_one('#id_email + .invalid-feedback').get_text())

    def test_4_3_4(self):
        '''更新処理（異常系：メールアドレス必須チェック）'''
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='テスト株式会社',
            email='old@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name_kana': 'テスト株式会社',
            'partner_type': 'customer',
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner.refresh_from_db()
        self.assertEqual(partner.email, 'old@example.com')

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('この項目は必須です。', soup.select_one('#id_email + .invalid-feedback').get_text())

    def test_4_3_5(self):
        '''更新処理（異常系：取引先区分不正）'''
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='テスト株式会社',
            email='old@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'invalid',
            'email': 'test@example.com',
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner.refresh_from_db()
        self.assertEqual(partner.partner_type, 'customer')

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertEqual('正しく選択してください。 invalid は候補にありません。', soup.select_one('#id_partner_type + .invalid-feedback').get_text())

    def test_4_3_6(self):
        '''更新処理（異常系：同一テナント内で取引先名称＋メールアドレス重複）'''
        # 既存データ1
        partner1 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='テスト株式会社',
            email='test@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        # 更新対象
        partner2 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='別会社',
            email='test2@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:update', args=[partner2.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com', 
        }

        # レスポンス確認
        response = self.client.post(url, data, HTTP_X_REQUESTED_WITH='XMLHttpRequest')
        self.assertEqual(response.status_code, 200)
        res_json = json.loads(response.content)
        self.assertFalse(res_json['success'])

        # 更新されていないこと
        partner2.refresh_from_db()
        self.assertEqual(partner2.partner_name, '別会社')  # 更新されていない

        # エラーメッセージ確認
        soup = BeautifulSoup(res_json['html'], 'html.parser')
        self.assertIn('同じ取引先名称とメールアドレスの組み合わせが既に登録されています。', soup.select_one('#id_partner_name + .invalid-feedback').get_text())
        self.assertIn('同じ取引先名称とメールアドレスの組み合わせが既に登録されています。', soup.select_one('#id_email + .invalid-feedback').get_text())

    # -------------------
    # 削除処理
    # -------------------
    def test_5_1_1(self):
        '''
        削除処理（正常系）
        '''
        # 取引先のテストデータ
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='テスト株式会社',
            email='delete@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        # 受注テストデータ
        sales_order = SalesOrder.objects.create(
            tenant=self.user.tenant,
            partner=partner,
            sales_order_no='SO-001',
            sales_order_date='2025-09-30',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:delete', args=[partner.id])
        response = self.client.post(
            url,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )
        self.assertEqual(response.status_code, 200)
        res_json = response.json()
        self.assertTrue(res_json['success'])

        # 取引先の削除確認
        self.assertFalse(Partner.objects.filter(id=partner.id).exists())

        # 受注データの取引先がNullになっていること
        sales_order.refresh_from_db()
        self.assertIsNone(sales_order.partner)
        
    def test_5_2_1(self):
        '''
        一括削除（正常系）
        '''
        # 削除対象データ作成
        partner1 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象1',
            email='p1@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        partner2 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象2',
            email='p2@example.com', 
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        
        # 受注データ作成
        sales_order_1 = SalesOrder.objects.create(
            tenant=self.user.tenant,
            partner=partner1,
            sales_order_no='SO-001',
            sales_order_date='2025-09-30',
            create_user=self.user,
            update_user=self.user,
        )
        
        # 受注データ作成
        sales_order_2 = SalesOrder.objects.create(
            tenant=self.user.tenant,
            partner=partner2,
            sales_order_no='SO-002',
            sales_order_date='2025-09-30',
            create_user=self.user,
            update_user=self.user,
        )
        
        # 処理実行
        url = reverse('partner_mst:bulk_delete')
        response = self.client.post(
            url,
            {'ids': [partner1.id, partner2.id]},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        res_json = response.json()
        self.assertIn('件削除しました', res_json['message'])
        self.assertIn('2件削除しました', res_json['message'])

        # DB確認
        self.assertFalse(Partner.objects.filter(id=partner1.id).exists())
        self.assertFalse(Partner.objects.filter(id=partner2.id).exists())
        
        # 受注データの取引先がすべてNullになっていること
        sales_order_1.refresh_from_db()
        sales_order_2.refresh_from_db()
        self.assertIsNone(sales_order_1.partner)
        self.assertIsNone(sales_order_2.partner)

    def test_5_2_2(self):
        '''
        一括削除（異常系：指定なし）
        '''
        # 削除対象データ作成
        partner1 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象1',
            email='p1@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        partner2 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象2',
            email='p2@example.com', 
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )

        url = reverse('partner_mst:bulk_delete')
        response = self.client.post(url, {})  # ids なし

        # リダイレクト確認
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('partner_mst:list'))

        # メッセージがセットされていることを確認
        storage = list(messages.get_messages(response.wsgi_request))
        self.assertTrue(any('削除対象が選択されていません' in str(m) for m in storage))

        # DB確認（削除されていない）
        self.assertTrue(Partner.objects.filter(id=partner1.id).exists())
        self.assertTrue(Partner.objects.filter(id=partner2.id).exists())
        
    # -------------------
    # CSVインポート
    # -------------------
    def _make_csv_file(self, rows, header=None):
        '''
        テスト用CSVファイルの作成
        '''
        if header is None:
            fieldnames = [
                'partner_name', 'partner_name_kana', 'partner_type',
                'contact_name', 'email', 'tel_number', 'postal_code',
                'state', 'city', 'address', 'address2'
            ]
        else:
            fieldnames = header
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return io.BytesIO(output.getvalue().encode('utf-8'))

    def test_6_1_1(self):
        '''CSVインポート（正常系：1件）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        rows = [
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test1@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            }
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 200)

        res_json = json.loads(response.content)
        self.assertEqual('1件をインポートしました。', res_json['message'])
        self.assertEqual(Partner.objects.count(), 1)
        
        # 登録値の確認
        partner = Partner.objects.first()
        self.assertEqual(partner.partner_name, 'テスト株式会社')
        self.assertEqual(partner.partner_name_kana, 'テストカブシキガイシャ')
        self.assertEqual(partner.partner_type, 'customer')
        self.assertEqual(partner.contact_name, '山田太郎')
        self.assertEqual(partner.email, 'test1@example.com')
        self.assertEqual(partner.tel_number, '0312345678')
        self.assertEqual(partner.postal_code, '1000001')
        self.assertEqual(partner.state, '東京都')
        self.assertEqual(partner.city, '千代田区')
        self.assertEqual(partner.address, '丸の内1-1-1')
        self.assertEqual(partner.address2, 'ビル3F')
        self.assertEqual(partner.create_user, self.user)
        self.assertEqual(partner.update_user, self.user)
        self.assertEqual(partner.tenant, self.user.tenant)
        now = timezone.now()
        self.assertLessEqual(abs((partner.created_at - now).total_seconds()), 5)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)

    def test_6_1_2(self):
        '''CSVインポート（正常系：n件）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
            
        rows = [
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test1@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            {
                'partner_name': 'サンプル商事',
                'partner_name_kana': 'サンプルショウジ',
                'partner_type': 'supplier',
                'contact_name': '佐藤花子',
                'email': 'test2@example.com',
                'tel_number': '0459876543',
                'postal_code': '2200002',
                'state': '神奈川県',
                'city': '横浜市西区',
                'address': 'みなとみらい2-2-2',
                'address2': 'ランドタワー10F',
            },
            {
                'partner_name': 'グローバル合同会社',
                'partner_name_kana': 'グローバルゴウドウガイシャ',
                'partner_type': 'both',
                'contact_name': '鈴木一郎',
                'email': 'test3@example.com',
                'tel_number': '0521234567',
                'postal_code': '4500003',
                'state': '愛知県',
                'city': '名古屋市中村区',
                'address': '名駅3-3-3',
                'address2': 'サウスタワー5F',
            },
            {
                'partner_name': '未来産業株式会社',
                'partner_name_kana': 'ミライサンギョウカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '高橋次郎',
                'email': 'test4@example.com',
                'tel_number': '0665432109',
                'postal_code': '5300004',
                'state': '大阪府',
                'city': '大阪市北区',
                'address': '梅田4-4-4',
                'address2': 'グランビル8F',
            },
            {
                'partner_name': 'クリエイト有限会社',
                'partner_name_kana': 'クリエイトユウゲンガイシャ',
                'partner_type': 'supplier',
                'contact_name': '田中三郎',
                'email': 'test5@example.com',
                'tel_number': '0753456789',
                'postal_code': '6000005',
                'state': '京都府',
                'city': '京都市下京区',
                'address': '四条通5-5-5',
                'address2': '中央ビル2F',
            },
        ]

        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 200)

        res_json = json.loads(response.content)
        self.assertEqual('5件をインポートしました。', res_json['message'])
        self.assertEqual(Partner.objects.count(), 5)
        
        # 登録値の確認
        partners = Partner.objects.order_by('email')  # email で並べて rows と突き合わせやすくする
        for row, partner in zip(sorted(rows, key=lambda x: x['email']), partners):
            self.assertEqual(partner.partner_name, row['partner_name'])
            self.assertEqual(partner.partner_name_kana, row['partner_name_kana'])
            self.assertEqual(partner.partner_type, row['partner_type'])
            self.assertEqual(partner.contact_name, row['contact_name'])
            self.assertEqual(partner.email, row['email'])
            self.assertEqual(partner.tel_number, row['tel_number'])
            self.assertEqual(partner.postal_code, row['postal_code'])
            self.assertEqual(partner.state, row['state'])
            self.assertEqual(partner.city, row['city'])
            self.assertEqual(partner.address, row['address'])
            self.assertEqual(partner.address2, row['address2'])
            self.assertEqual(partner.create_user, self.user)
            self.assertEqual(partner.update_user, self.user)
            self.assertEqual(partner.tenant, self.user.tenant)
            
    def test_6_1_3(self):
        '''CSVインポート（正常系：0件）'''
        url = reverse('partner_mst:import_csv')

        # 取引先データを全削除
        Partner.objects.all().delete()

        # 空データ
        rows = []
        file = self._make_csv_file(rows)

        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 200)

        res_json = json.loads(response.content)
        self.assertEqual('0件をインポートしました。', res_json['message'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_1_4(self):
        '''CSVインポート（正常系：1件）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        partner_name = 'a' * 255
        partner_name_kana = 'b' * 255
        partner_type = 'customer'
        contact_name = 'c' * 255
        tel_number = '1' * 15
        email = 'test@example.com'
        postal_code = '1' * 7
        state = 'f' * 5
        city = 'g' * 255
        address = 'h' * 255
        address2 = 'i' * 255
        
        rows = [{
            'partner_name': partner_name,
            'partner_name_kana': partner_name_kana,
            'partner_type': partner_type,
            'contact_name': contact_name,
            'tel_number': tel_number,
            'email': email,
            'postal_code': postal_code,
            'state': state,
            'city': city,
            'address': address,
            'address2': address2,
        }]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 200)

        res_json = json.loads(response.content)
        self.assertEqual('1件をインポートしました。', res_json['message'])
        self.assertEqual(Partner.objects.count(), 1)
        
        # DB登録確認
        partner = Partner.objects.get(partner_name=partner_name)
        self.assertEqual(partner.partner_name, partner_name)
        self.assertEqual(partner.partner_name_kana, partner_name_kana)
        self.assertEqual(partner.partner_type, partner_type)
        self.assertEqual(partner.contact_name, contact_name)
        self.assertEqual(partner.tel_number, tel_number)
        self.assertEqual(partner.email, email)
        self.assertEqual(partner.postal_code, postal_code)
        self.assertEqual(partner.state, state)
        self.assertEqual(partner.city, city)
        self.assertEqual(partner.address, address)
        self.assertEqual(partner.address2, address2)
        self.assertEqual(partner.is_deleted, False)
        self.assertEqual(partner.create_user, self.user)
        self.assertEqual(partner.update_user, self.user)
        self.assertEqual(partner.tenant, self.user.tenant)
        now = timezone.now()
        self.assertLessEqual(abs((partner.created_at - now).total_seconds()), 5)
        self.assertLessEqual(abs((partner.updated_at - now).total_seconds()), 5)


    def test_6_2_1(self):
        '''CSVインポート（異常系：ヘッダ不正）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # 不正なヘッダを作成
        fieldnames = [
            'partner_name_test', 'partner_name_kana', 'partner_type',
            'contact_name', 'email', 'tel_number', 'postal_code',
            'state', 'city', 'address', 'address2'
        ]
        
        rows = [
            {
                'partner_name_test': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test1@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            }
        ]
        file = self._make_csv_file(rows, header=fieldnames)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVヘッダが正しくありません。', res_json['error'])
        self.assertEqual(f"期待: {DATA_COLUMNS}, 実際: ['partner_name_test', 'partner_name_kana', 'partner_type', 'contact_name', 'email', 'tel_number', 'postal_code', 'state', 'city', 'address', 'address2']", res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_2(self):
        '''CSVインポート（異常系：取引先名称バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 取引先名称：必須エラー
            {
                'partner_name': '',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            # 取引先名称：文字数超過エラー
            {
                'partner_name': 'a' * 256,
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: partner_name: この項目は必須です。', '3行目: partner_name: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_3(self):
        '''CSVインポート（異常系：取引先名称（カナ）バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 取引先名称（カナ）：文字数超過エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'a' * 256,
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test1@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: partner_name_kana: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_4(self):
        '''CSVインポート（異常系：取引先区分バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 取引先区分：必須エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': '',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            # 取引先区分：不正値エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'invalid',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: partner_type: この項目は必須です。', '3行目: partner_type: 正しく選択してください。 invalid は候補にありません。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_5(self):
        '''CSVインポート（異常系：担当者名バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 担当者名：文字数超過エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': 'a' * 256,
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: contact_name: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_6(self):
        '''CSVインポート（異常系：電話番号バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 電話番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '1'*16,
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            # 電話番号：不正値
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': 'test',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: tel_number: この値は 15 文字以下でなければなりません( 16 文字になっています)。', '3行目: tel_number: 半角数字のみを入力してください。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_7(self):
        '''CSVインポート（異常系：メールアドレスバリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # メールアドレス：必須エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': '',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            # メールアドレス不正エラー
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test@@test',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: email: この項目は必須です。', '3行目: email: 有効なメールアドレスを入力してください。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_8(self):
        '''CSVインポート（異常系：郵便番号バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 郵便番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '1'*16,
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
            # 郵便番号：不正値
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': 'test',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: tel_number: この値は 15 文字以下でなければなりません( 16 文字になっています)。', '3行目: tel_number: 半角数字のみを入力してください。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_9(self):
        '''CSVインポート（異常系：都道府県バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 郵便番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': 'a'*6,
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: state: この値は 5 文字以下でなければなりません( 6 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_10(self):
        '''CSVインポート（異常系：市町村区バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 郵便番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': 'a'*256,
                'address': '丸の内1-1-1',
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: city: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_11(self):
        '''CSVインポート（異常系：住所バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 郵便番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': 'a'*256,
                'address2': 'ビル3F',
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: address: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_6_2_12(self):
        '''CSVインポート（異常系：住所2バリデーションエラー）'''
        url = reverse('partner_mst:import_csv')
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # データ作成
        rows = [
            # 郵便番号：文字数超過
            {
                'partner_name': 'テスト株式会社',
                'partner_name_kana': 'テストカブシキガイシャ',
                'partner_type': 'customer',
                'contact_name': '山田太郎',
                'email': 'test2@example.com',
                'tel_number': '0312345678',
                'postal_code': '1000001',
                'state': '東京都',
                'city': '千代田区',
                'address': '丸の内1-1-1',
                'address2': 'a'*256,
            },
        ]
        file = self._make_csv_file(rows)
        response = self.client.post(url, {'file': file})
        self.assertEqual(response.status_code, 400)

        res_json = json.loads(response.content)
        self.assertEqual('CSVに問題があります。', res_json['error'])
        self.assertEqual(['2行目: address2: この値は 255 文字以下でなければなりません( 256 文字になっています)。'], res_json['details'])
        self.assertEqual(Partner.objects.count(), 0)

    # -------------------
    # CSVエクスポート
    # -------------------
    def _request_and_parse(self, url: str):
        '''エクスポートを叩いてレスポンスとCSV行を返す共通処理'''
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        content = response.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        # ヘッダ確認
        self.assertEqual(rows[0], DATA_COLUMNS)
        return rows

    def _assert_csv_matches_queryset(self, rows, queryset):
        '''CSVの行とクエリセットを突き合わせ'''
        for row, partner in zip(rows[1:], queryset):
            expected = [
                str(getattr(partner, col)) if getattr(partner, col) is not None else ''
                for col in DATA_COLUMNS
            ]
            self.assertEqual(row, expected)

    def test_7_1_1(self):
        '''CSVエクスポート（正常系：n件）'''
        url = reverse('partner_mst:export_csv')
        rows = self._request_and_parse(url)
        # 件数チェック（事前に8件のデータがある想定）
        self.assertEqual(len(rows) - 1, 8)
        partners = Partner.objects.filter(
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_2(self):
        '''CSVエクスポート（正常系：0件）'''
        Partner.objects.all().delete()
        url = reverse('partner_mst:export_csv')
        rows = self._request_and_parse(url)
        # ヘッダのみ
        self.assertEqual(len(rows), 1)

    def test_7_1_3(self):
        '''CSVエクスポート（正常系：取引先名称検索）'''
        target = '株式会社アルファ'
        url = reverse('partner_mst:export_csv') + f'?search_partner_name={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            partner_name__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_4(self):
        '''CSVエクスポート（正常系：区分検索）'''
        target = 'both'
        url = reverse('partner_mst:export_csv') + f'?search_partner_type={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            partner_type__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_5(self):
        '''CSVエクスポート（正常系：担当者名検索）'''
        target = '佐藤'
        url = reverse('partner_mst:export_csv') + f'?search_contact_name={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            contact_name__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_6(self):
        '''CSVエクスポート（正常系：電話番号検索）'''
        target = '0356781234'
        url = reverse('partner_mst:export_csv') + f'?search_tel_number={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            tel_number__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_7(self):
        '''CSVエクスポート（正常系：メールアドレス検索）'''
        target = 'cs.'
        url = reverse('partner_mst:export_csv') + f'?search_email={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            email__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_8(self):
        '''CSVエクスポート（正常系：都道府県検索）'''
        target = '北海道'
        url = reverse('partner_mst:export_csv') + f'?search_address={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            state__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_9(self):
        '''CSVエクスポート（正常系：市区町村検索）'''
        target = '千代田区'
        url = reverse('partner_mst:export_csv') + f'?search_address={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            city__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_10(self):
        '''CSVエクスポート（正常系：住所検索）'''
        target = 'フーズビル'
        url = reverse('partner_mst:export_csv') + f'?search_address={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            address__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_7_1_11(self):
        '''CSVエクスポート（正常系：住所2検索）'''
        target = 'セカンドアドレス'
        url = reverse('partner_mst:export_csv') + f'?search_address={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            address2__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(len(rows) - 1, partners.count())
        self._assert_csv_matches_queryset(rows, partners)

    def test_export_with_nottest_7_1_12_found(self):
        '''CSVエクスポート（正常系：存在しない名称 → 0件）'''
        target = 'グローバル貿易株式会社'
        url = reverse('partner_mst:export_csv') + f'?search_partner_name={target}'
        rows = self._request_and_parse(url)
        partners = Partner.objects.filter(
            partner_name__icontains=target,
            tenant=self.user.tenant,
            is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(partners.count(), 0)
        self.assertEqual(len(rows) - 1, 0)

            
    # -------------------
    # EXCELエクスポート
    # -------------------
    def _get_excel_rows(self, response):
        '''Excelレスポンスをopenpyxlで読み取り、行データを返す'''
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        return [[str(cell.value) if cell.value is not None else '' for cell in row]
                for row in ws.iter_rows()]

    def _assert_export(self, url, expected_count, queryset):
        '''共通チェック: Excelを取得→件数とDB値を照合'''
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        rows = self._get_excel_rows(response)

        # ヘッダ確認
        self.assertEqual(rows[0], DATA_COLUMNS)

        # 件数確認
        self.assertEqual(len(rows) - 1, expected_count)

        # データ内容確認
        for row, partner in zip(rows[1:], queryset):
            expected = [
                str(getattr(partner, col)) if getattr(partner, col) is not None else ''
                for col in DATA_COLUMNS
            ]
            self.assertEqual(row, expected)

    def test_8_1_1(self):
        '''Excelエクスポート（正常系：n件）'''
        qs = Partner.objects.filter(
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(reverse('partner_mst:export_excel'), qs.count(), qs)

    def test_8_1_2(self):
        '''Excelエクスポート（正常系：0件）'''
        Partner.objects.all().delete()
        qs = Partner.objects.none()
        self._assert_export(reverse('partner_mst:export_excel'), 0, qs)

    def test_8_1_3(self):
        '''Excelエクスポート（正常系：取引先名称検索）'''
        target = '株式会社アルファ'
        qs = Partner.objects.filter(
            partner_name__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_partner_name={target}',
            qs.count(), qs
        )

    def test_8_1_4(self):
        '''Excelエクスポート（正常系：取引先区分検索）'''
        target = 'both'
        qs = Partner.objects.filter(
            partner_type__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_partner_type={target}',
            qs.count(), qs
        )

    def test_8_1_5(self):
        '''Excelエクスポート（正常系：担当者名検索）'''
        target = '佐藤'
        qs = Partner.objects.filter(
            contact_name__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_contact_name={target}',
            qs.count(), qs
        )

    def test_8_1_6(self):
        '''Excelエクスポート（正常系：電話番号検索）'''
        target = '0356781234'
        qs = Partner.objects.filter(
            tel_number__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_tel_number={target}',
            qs.count(), qs
        )

    def test_8_1_7(self):
        '''Excelエクスポート（正常系：メールアドレス検索）'''
        target = 'cs.'
        qs = Partner.objects.filter(
            email__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_email={target}',
            qs.count(), qs
        )

    def test_8_1_8(self):
        '''Excelエクスポート（正常系：都道府県検索）'''
        target = '北海道'
        qs = Partner.objects.filter(
            state__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_address={target}',
            qs.count(), qs
        )

    def test_8_1_9(self):
        '''Excelエクスポート（正常系：市区町村検索）'''
        target = '千代田区'
        qs = Partner.objects.filter(
            city__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_address={target}',
            qs.count(), qs
        )

    def test_8_1_10(self):
        '''Excelエクスポート（正常系：住所検索）'''
        target = 'フーズビル'
        qs = Partner.objects.filter(
            address__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_address={target}',
            qs.count(), qs
        )

    def test_8_1_11(self):
        '''Excelエクスポート（正常系：住所2検索）'''
        target = 'セカンドアドレス'
        qs = Partner.objects.filter(
            address2__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_address={target}',
            qs.count(), qs
        )

    def test_8_1_12(self):
        '''Excelエクスポート（正常系：存在しない取引先名称検索 → 0件）'''
        target = 'グローバル貿易株式会社'
        qs = Partner.objects.filter(
            partner_name__icontains=target,
            tenant=self.user.tenant, is_deleted=False
        ).order_by('partner_name')
        self.assertEqual(qs.count(), 0)
        self._assert_export(
            reverse('partner_mst:export_excel') + f'?search_partner_name={target}',
            0, qs
        )

    # -------------------
    # 直リンク禁止
    # -------------------
    def test_9_1_1(self):
        '''一覧画面表示（異常系：未ログイン）'''
        url = reverse('partner_mst:list')

        # 未ログイン状態でアクセス
        self.client.logout()
        response = self.client.get(url, follow=False)

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, '/login/?next=/partner_mst/')
        
    def test_9_2_1(self):
        '''登録画面画面表示（異常系：未ログイン）'''
        url = reverse('partner_mst:create')

        # 未ログイン状態でアクセス
        self.client.logout()
        response = self.client.get(url, follow=False)

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, '/login/?next=/partner_mst/create/')
        
    def test_9_2_2(self):
        '''登録画処理（異常系：未ログイン）'''
        url = reverse('partner_mst:create')

        # 未ログイン状態でアクセス
        self.client.logout()
        response = self.client.get(url, follow=False)
        
        # 取引先データを全削除
        Partner.objects.all().delete()
        
        # 登録データ作成
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com',
        }

        # 処理実行
        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, '/login/?next=/partner_mst/create/')
        
        # DB登録されていないことを確認
        self.assertEqual(Partner.objects.count(), 0)
        
    def test_9_3_1(self):
        '''更新画面表示（異常系：未ログイン）'''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='ログイン株式会社',
            partner_type='supplier',
            email='login@example.com',
            create_user=self.user,
            update_user=self.user,
        )
        
        url = reverse('partner_mst:update', args=[partner.id])

        # 未ログイン状態でアクセス
        self.client.logout()
        response = self.client.get(url, follow=False)

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, f'/login/?next=/partner_mst/{partner.id}/update/')
        
    def test_9_3_2(self):
        '''更新処理（異常系：未ログイン）'''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='ログイン株式会社',
            partner_type='supplier',
            email='login@example.com',
            create_user=self.user,
            update_user=self.user,
        )
        
        # 更新処理のURLとデータ作成
        url = reverse('partner_mst:update', args=[partner.id])
        data = {
            'partner_name': 'テスト株式会社',
            'partner_type': 'customer',
            'email': 'test@example.com',
        }

        # ログアウト
        self.client.logout()
        
        # 処理実行
        response = self.client.post(
            url,
            data,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, f'/login/?next=/partner_mst/{partner.id}/update/')
        
        # データ更新されていないことを確認
        partner.refresh_from_db()
        self.assertEqual(partner.partner_name, 'ログイン株式会社')
        self.assertEqual(partner.partner_type, 'supplier')
        self.assertEqual(partner.email, 'login@example.com')
        
    def test_9_4_1(self):
        '''削除処理（異常系：未ログイン）'''
        # 事前データ作成
        partner = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='ログイン株式会社',
            partner_type='supplier',
            email='login@example.com',
            create_user=self.user,
            update_user=self.user,
        )
        
        # 更新処理のURLとデータ作成
        url = reverse('partner_mst:delete', args=[partner.id])

        # ログアウト
        self.client.logout()
        
        # 処理実行
        response = self.client.post(
            url,
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, f'/login/?next=/partner_mst/{partner.id}/delete/')
        
        # データ削除されていないことを確認
        self.assertTrue(Partner.objects.filter(id=partner.id).exists())
        
    def test_9_5_1(self):
        '''
        一括削除（異常系：未ログイン）
        '''
        # 削除対象データ作成
        partner1 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象1',
            email='p1@example.com',
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        partner2 = Partner.objects.create(
            tenant=self.user.tenant,
            partner_name='削除対象2',
            email='p2@example.com', 
            partner_type='customer',
            create_user=self.user,
            update_user=self.user,
        )
        
        # ログアウト
        self.client.logout()

        # 処理実行
        url = reverse('partner_mst:bulk_delete')
        response = self.client.post(
            url,
            {'ids': [partner1.id, partner2.id]},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest'
        )

        # ステータスコード確認
        self.assertEqual(response.status_code, 302)
        
        # 遷移後の画面確認
        self.assertRedirects(response, '/login/?next=/partner_mst/bulk_delete/')

        # DB確認
        self.assertTrue(Partner.objects.filter(id=partner1.id).exists())
        self.assertTrue(Partner.objects.filter(id=partner2.id).exists())

    def test_9_6_1(self):
        '''CSVインポート（異常系：未ログイン）'''
        url = reverse('partner_mst:import_csv')

        # 未ログイン状態にする
        self.client.logout()

        # アクセス（リダイレクトされるはず）
        response = self.client.get(url, follow=False)

        # リダイレクトを確認
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, '/login/?next=/partner_mst/import/csv')

    def test_9_7_1(self):
        '''CSVエクスポート（異常系：未ログイン）'''
        url = reverse('partner_mst:export_csv')

        # 未ログイン状態にする
        self.client.logout()

        # アクセス（リダイレクトされるはず）
        response = self.client.get(url, follow=False)

        # リダイレクトを確認
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, '/login/?next=/partner_mst/export/csv')

    def test_9_7_1(self):
        '''EXCELエクスポート（異常系：未ログイン）'''
        url = reverse('partner_mst:export_excel')

        # 未ログイン状態にする
        self.client.logout()

        # アクセス（リダイレクトされるはず）
        response = self.client.get(url, follow=False)

        # リダイレクトを確認
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, '/login/?next=/partner_mst/export/excel')

