import csv

from django.conf import settings
from django.http import JsonResponse
from django.views import View
from django.db import transaction, IntegrityError
from datetime import datetime
from django.http import HttpResponse
from django.db import models
from django.http import HttpResponseForbidden
from register.constants import PRIVILEGE_EDITOR, PRIVILEGE_MANAGER, PRIVILEGE_SYSTEM
from django import forms
import openpyxl


class ExcelExportBaseView(View):
    model_class = None  # 出力対象のモデルクラス
    filename_prefix = 'export'  # サブクラスで指定
    headers: list[str] = []  # 出力ヘッダ定義

    def get(self, request, *args, **kwargs):
        # ファイル名設定
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'{self.filename_prefix}_{timestamp}.xlsx'

        # データ取得
        data = self.get_queryset(request).filter(is_deleted=False, tenant=request.user.tenant)

        # Workbook 作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'data'

        # ヘッダ書き込み
        for col, header in enumerate(self.headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # データ書き込み
        for row_idx, rec in enumerate(data, start=2):
            values = self.row(rec)
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # レスポンス作成
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{file_name}"

        wb.save(response)
        return response

    def get_queryset(self, request):
        return self.model_class.objects.all()

    def row(self, rec):
        raise NotImplementedError('サブクラスで row() を実装してください')


class CSVExportBaseView(View):
    model_class = None  # 出力対象のモデルクラス
    filename_prefix = 'export'  # サブクラスで指定
    headers: list[str] = []  # 出力ヘッダ定義

    def get(self, request, *args, **kwargs):
        # ファイル名設定
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f'{self.filename_prefix}_{timestamp}.csv'

        # データ取得
        data = self.get_queryset(request)

        # レスポンス準備
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{file_name}"

        writer = csv.writer(response)

        # ヘッダ出力
        writer.writerow(self.headers)

        # データ出力
        for rec in data:
            writer.writerow(self.row(rec))

        return response

    def get_queryset(self, request):
        '''サブクラスで必要に応じてフィルタリングを行う'''
        return self.model_class.objects.filter(is_deleted=False, tenant=request.user.tenant)

    def row(self, rec):
        '''サブクラスで1行分のリストを返す'''
        raise NotImplementedError('サブクラスで row() を実装してください')


class CSVImportBaseView(View):
    '''
    CSV Import機能の基底クラス（日本語・英語ヘッダ両対応）
    expected_headers: 日本語ヘッダのリスト
    HEADER_MAP: 日本語→英語の辞書
    '''
    expected_headers: list[str] = []
    model_class = None
    unique_field = None
    HEADER_MAP = {}

    # ------------------------------------------------------------
    # 内部ユーティリティ：エラーを日本語ラベルに変換
    # ------------------------------------------------------------
    def _format_errors_with_verbose_name(self, form):
        '''
        form.errors を日本語フィールド名付きに変換して返す
        '''
        messages = []
        for field, errs in form.errors.items():
            try:
                # フィールド定義に label（フォーム）or verbose_name（モデル）があれば使う
                if hasattr(form.fields[field], 'label'):
                    label = form.fields[field].label
                else:
                    label = self.model_class._meta.get_field(field).verbose_name
            except Exception:
                # 万一取得できなければ英語フィールド名のまま
                label = field
            messages.append(f"{label}: {'; '.join(errs)}")
        return ' / '.join(messages)

    # ------------------------------------------------------------
    # POST: CSVインポート処理
    # ------------------------------------------------------------
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'error': 'ファイルが選択されていません'}, status=400)

        # ------------------------------------------------------------
        # ファイルサイズチェック
        # ------------------------------------------------------------
        if file.size > settings.MAX_FILE_SIZE:
            return JsonResponse({'error': 'ファイルサイズが上限を超えています。'}, status=400)

        # ------------------------------------------------------------
        # CSV読み込み
        # ------------------------------------------------------------
        file_data = file.read()
        try:
            decoded_data = file_data.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded_data = file_data.decode('cp932')

        decoded_file = decoded_data.splitlines()
        reader = csv.DictReader(decoded_file)

        # ------------------------------------------------------------
        # ヘッダチェック（日本語・英語両対応）
        # ------------------------------------------------------------
        if reader.fieldnames is None:
            return JsonResponse({'error': 'CSVにヘッダ行が存在しません。'}, status=400)

        normalize = lambda x: x.strip().replace('　', '')
        normalized_actual = [normalize(h) for h in reader.fieldnames]
        normalized_expected = [normalize(h) for h in self.expected_headers]

        allowed_headers = set(normalized_expected) | set(self.HEADER_MAP.values())

        # ヘッダ欠落パターン
        missing = [h for h in normalized_expected
                   if h not in normalized_actual and self.HEADER_MAP.get(h) not in normalized_actual]
        if missing:
            return JsonResponse({
                'error': 'CSVヘッダが正しくありません。',
                'details': f'不足: {missing} / 期待: {self.expected_headers} / 実際: {reader.fieldnames}'
            }, status=400)

        # 不要ヘッダパターン
        unexpected = [h for h in normalized_actual if h not in allowed_headers]
        if unexpected:
            return JsonResponse({
                'error': '想定外のヘッダが含まれています。',
                'details': f'不要: {unexpected} / 許可: {list(allowed_headers)}'
            }, status=400)

        # ヘッダ重複パターン
        duplicates = [h for h in set(normalized_actual) if normalized_actual.count(h) > 1]
        if duplicates:
            return JsonResponse({
                'error': 'CSVヘッダが重複しています。',
                'details': f'重複: {duplicates}',
            }, status=400)


        # ------------------------------------------------------------
        # 重複チェックデータ準備
        # ------------------------------------------------------------
        if isinstance(self.unique_field, (list, tuple)):
            existing = set(self.model_class.objects.values_list(*self.unique_field))
        else:
            existing = set(self.model_class.objects.values_list(self.unique_field, flat=True))

        # ------------------------------------------------------------
        # 行ごとのバリデーション
        # ------------------------------------------------------------
        objects_to_create, errors = [], []
        for idx, row in enumerate(reader, start=2):
            normalized_row = {}
            for key, value in row.items():
                key_norm = normalize(key)

                # 日本語ヘッダ → 英語キーに変換
                if key_norm in self.HEADER_MAP:
                    normalized_row[self.HEADER_MAP[key_norm]] = value

                # 英語ヘッダ → そのまま使う
                elif key_norm in self.HEADER_MAP.values():
                    normalized_row[key_norm] = value
                else:
                    continue

            # --------------------------------------------------------
            # 各行ごとに validate_row() 実行
            # --------------------------------------------------------
            try:
                obj, err = self.validate_row(
                    row=normalized_row, idx=idx, existing=existing, request=request
                )
            except Exception as e:
                # validate_row 内部で例外が出ても1行目などで特定できるように補足
                errors.append(f"{idx}行目: {str(e)}")
                continue

            # validate_row 側が form.is_valid() のエラーを返した場合
            if isinstance(err, forms.ModelForm):
                msg = self._format_errors_with_verbose_name(err)
                errors.append(f"{idx}行目: {msg}")
            elif err:
                errors.append(err)
            elif obj:
                objects_to_create.append(obj)

        # ------------------------------------------------------------
        # エラーがあればJSONで返す
        # ------------------------------------------------------------
        if errors:
            return JsonResponse({'error': 'CSVに問題があります。', 'details': errors}, status=400)

        # ------------------------------------------------------------
        # 登録実行
        # ------------------------------------------------------------
        try:
            with transaction.atomic():
                self.model_class.objects.bulk_create(objects_to_create)
        except IntegrityError as e:
            return JsonResponse({'error': '登録中にDBエラーが発生しました。', 'details': [str(e)]}, status=500)

        return JsonResponse({'message': f'{len(objects_to_create)}件をインポートしました。'})

    # ------------------------------------------------------------
    # サブクラスで実装すべき：1行単位の検証ロジック
    # ------------------------------------------------------------
    def validate_row(self, row: dict, idx: int, existing: set, request):
        '''
        各CSV行を検証して、成功なら (obj, None)、失敗なら (None, form or str) を返す
        '''
        raise NotImplementedError


class BaseModel(models.Model):
    '''
    共通基底クラス
    - 論理削除
    - 作成日時 / 更新日時
    - 作成者 / 更新者
    '''

    tenant = models.ForeignKey('tenant_mst.Tenant', on_delete=models.CASCADE, related_name="%(class)ss", null=False, blank=False, verbose_name='所属テナント')
    is_deleted = models.BooleanField(default=False, verbose_name='削除フラグ')
    created_at = models.DateTimeField(auto_now_add=True, verbose_name='作成日時')
    updated_at = models.DateTimeField(auto_now=True, verbose_name='更新日時')
    create_user = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        related_name="%(class)s_creator",
        verbose_name='作成者'
    )
    update_user = models.ForeignKey(
        settings.AUTH_USER_MODEL,
        on_delete=models.SET_NULL,
        null=True,
        related_name="%(class)s_updater",
        verbose_name='更新者'
    )

    class Meta:
        abstract = True

class PrivilegeRequiredMixin():
    '''
    参照権限以下のユーザーのアクセスを制限する
    '''
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if int(request.user.privilege) > int(PRIVILEGE_EDITOR):
            return HttpResponseForbidden('アクセス権限がありません')
        return super().dispatch(request, *args, **kwargs)

class ManagerOverMixin():
    '''
    更新権限以下のユーザーのアクセスを制限する
    '''
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if int(request.user.privilege)  > int(PRIVILEGE_MANAGER):
            return HttpResponseForbidden('アクセス権限がありません。')
        return super().dispatch(request, *args, **kwargs)

class SystemUserOnlyMixin():
    '''
    システム管理者のみアクセス可能
    '''
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if int(request.user.privilege)  != int(PRIVILEGE_SYSTEM):
            return HttpResponseForbidden('アクセス権限がありません。')
        return super().dispatch(request, *args, **kwargs)
