import io
import csv
from django.test import TestCase, Client
from django.urls import reverse
from datetime import date
from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from product_mst.models import Product, ProductCategory
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook


class ProductViewTests(TestCase):
    def setUp(self):
        '''共通データ作成'''
        self.client = Client()
        self.user = get_user_model().objects.create_user(
            username='tester', password='password123'
        )
        self.client.login(username='tester', password='password123')

        self.category = ProductCategory.objects.create(product_category_name='文房具')
        self.product = Product.objects.create(
            product_code='PRD001',
            product_name='テスト商品',
            price=100,
            description='テスト説明文',
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            product_category=self.category,
            create_user=self.user,
            update_user=self.user,
        )

    # -------------------
    # CRUD (通常)
    # -------------------

    def test_product_list_view(self):
        '''商品一覧が表示される'''
        url = reverse('product_mst:product_list')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'テスト商品')
        
    def test_product_list_view_without_login(self):
        """未ログイン時に商品一覧へアクセス → ログインページへリダイレクト"""
        # セッションを強制的にクリアして未ログイン状態にする
        self.client.cookies.clear()

        url = reverse('product_mst:product_list')
        response = self.client.get(url)

        # LoginRequiredMixin が効いていればログインページへリダイレクト
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)

        # ログインページにリダイレクトされることを確認
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)

    def test_product_create_view_and_flash(self):
        '''商品登録 + フラッシュメッセージ + 登録内容確認'''
        url = reverse('product_mst:product_create')
        response = self.client.post(
            url,
            {
                'product_code': 'PRD002',
                'product_name': '新商品',
                'start_date': '2025-01-01',
                'end_date': '2025-12-31',
                'product_category': self.category.id,
                'price': 500,
                'description': 'テスト説明',
            },
            follow=True,
        )

        # レスポンスが正常か
        self.assertEqual(response.status_code, 200)

        # 登録された商品を取得
        product = Product.objects.get(product_code='PRD002')

        # 各フィールドの内容を確認
        self.assertEqual(product.product_name, '新商品')
        self.assertEqual(product.start_date.strftime('%Y-%m-%d'), '2025-01-01')
        self.assertEqual(product.end_date.strftime('%Y-%m-%d'), '2025-12-31')
        self.assertEqual(product.product_category, self.category)
        self.assertEqual(product.price, 500)
        self.assertEqual(product.description, 'テスト説明')
        self.assertEqual(product.create_user, self.user)
        self.assertEqual(product.update_user, self.user)

        # フラッシュメッセージを確認
        messages = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('新商品' in m for m in messages))
        self.assertTrue(any('登録' in m for m in messages))


    def test_product_update_view_and_flash(self):
        '''商品更新 + フラッシュメッセージ + 更新内容確認'''
        url = reverse('product_mst:product_update', kwargs={'pk': self.product.pk})
        response = self.client.post(
            url,
            {
                'product_code': 'PRD001',
                'product_name': '商品updated',
                'start_date': '2025-02-01',
                'end_date': '2026-01-01',
                'product_category': self.category.id,
                'price': 999,
                'description': '更新後の説明',
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        # DB再読み込み
        self.product.refresh_from_db()

        # 各フィールドの内容を確認
        self.assertEqual(self.product.product_code, 'PRD001')
        self.assertEqual(self.product.product_name, '商品updated')
        self.assertEqual(self.product.start_date.strftime('%Y-%m-%d'), '2025-02-01')
        self.assertEqual(self.product.end_date.strftime('%Y-%m-%d'), '2026-01-01')
        self.assertEqual(self.product.product_category, self.category)
        self.assertEqual(self.product.price, 999)
        self.assertEqual(self.product.description, '更新後の説明')
        self.assertEqual(self.product.create_user, self.user)   # 作成者は変わらない
        self.assertEqual(self.product.update_user, self.user)   # 更新者はログインユーザーに上書き

        # フラッシュメッセージを確認
        messages = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('商品updated' in m for m in messages))
        self.assertTrue(any('更新' in m for m in messages))


    def test_product_delete_view_and_flash(self):
        '''商品削除（論理削除） + フラッシュメッセージ'''
        url = reverse('product_mst:product_delete', kwargs={'pk': self.product.pk})
        response = self.client.post(url, follow=True)
        self.assertEqual(response.status_code, 200)
        deleted_product = Product.objects.get(pk=self.product.pk)
        self.assertTrue(deleted_product.is_deleted)
        messages = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('テスト商品' in m for m in messages))
        self.assertTrue(any('削除' in m for m in messages))

    # -------------------
    # CRUD (モーダル)
    # -------------------

    def test_product_create_modal_post_success_with_flash(self):
        '''モーダル登録成功 + フラッシュメッセージ + 登録内容確認'''
        url = reverse('product_mst:product_create_modal')
        response = self.client.post(
            url,
            {
                'product_code': 'NEW100',
                'product_name': 'モーダル新商品',
                'start_date': '2025-01-01',
                'end_date': '2025-12-31',
                'product_category': self.category.id,
                'price': 100,
                'description': 'モーダル経由で登録',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
            follow=True,
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])

        # 登録された商品を取得
        product = Product.objects.get(product_code='NEW100')

        # 各フィールドの内容を確認
        self.assertEqual(product.product_code, 'NEW100')
        self.assertEqual(product.product_name, 'モーダル新商品')
        self.assertEqual(product.start_date.strftime('%Y-%m-%d'), '2025-01-01')
        self.assertEqual(product.end_date.strftime('%Y-%m-%d'), '2025-12-31')
        self.assertEqual(product.product_category, self.category)
        self.assertEqual(product.price, 100)
        self.assertEqual(product.description, 'モーダル経由で登録')
        self.assertEqual(product.create_user, self.user)
        self.assertEqual(product.update_user, self.user)

        # フラッシュメッセージ確認
        messages = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('モーダル新商品' in m for m in messages))
        self.assertTrue(any('登録' in m for m in messages))


    def test_product_create_modal_post_invalid_required(self):
        '''モーダル登録 異常系（必須エラー: product_code 空）'''
        url = reverse('product_mst:product_create_modal')
        response = self.client.post(
            url,
            {
                'product_code': '',
                'product_name': '不正商品',
                'start_date': '2025-01-01',
                'end_date': '2025-12-31',
                'product_category': self.category.id,
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['success'])
        self.assertFalse(Product.objects.filter(product_name='不正商品').exists())

    def test_product_update_modal_post_success_with_flash(self):
        '''モーダル更新成功 + フラッシュメッセージ + 更新内容確認'''
        url = reverse('product_mst:product_update_modal', kwargs={'pk': self.product.pk})
        response = self.client.post(
            url,
            {
                'product_code': 'PRD001',
                'product_name': 'モーダル更新',
                'start_date': '2025-02-01',
                'end_date': '2026-01-01',
                'product_category': self.category.id,
                'price': 777,
                'description': 'モーダルで更新しました',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
            follow=True,
        )

        # レスポンス確認
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])

        # DB再読み込みしてフィールドを確認
        self.product.refresh_from_db()
        self.assertEqual(self.product.product_code, 'PRD001')
        self.assertEqual(self.product.product_name, 'モーダル更新')
        self.assertEqual(self.product.start_date.strftime('%Y-%m-%d'), '2025-02-01')
        self.assertEqual(self.product.end_date.strftime('%Y-%m-%d'), '2026-01-01')
        self.assertEqual(self.product.product_category, self.category)
        self.assertEqual(self.product.price, 777)
        self.assertEqual(self.product.description, 'モーダルで更新しました')
        self.assertEqual(self.product.create_user, self.user)   # 作成者は変わらない
        self.assertEqual(self.product.update_user, self.user)   # 更新者は上書きされる

        # フラッシュメッセージ確認
        messages = [m.message for m in get_messages(response.wsgi_request)]
        self.assertTrue(any('モーダル更新' in m for m in messages))
        self.assertTrue(any('更新' in m for m in messages))


    def test_product_update_modal_post_invalid_date(self):
        '''モーダル更新 異常系（終了日 < 開始日）'''
        url = reverse('product_mst:product_update_modal', kwargs={'pk': self.product.pk})
        response = self.client.post(
            url,
            {
                'product_code': 'PRD001',
                'product_name': '不正更新',
                'start_date': '2025-01-10',
                'end_date': '2025-01-05',
                'product_category': self.category.id,
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['success'])
        self.product.refresh_from_db()
        self.assertEqual(self.product.product_name, 'テスト商品')

    # -------------------
    # Export
    # -------------------

    def test_export_csv_content_exact(self):
        '''CSVエクスポートの中身を完全一致で確認'''
        url = reverse('product_mst:export_csv')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')

        content = response.content.decode('utf-8').splitlines()
        reader = csv.reader(content)
        rows = list(reader)

        expected_header = [
            'product_code', 'product_name', 'start_date', 'end_date',
            'product_category', 'price', 'description',
        ] + [col for col in rows[0][7:]]
        self.assertEqual(rows[0], expected_header)

        expected_row = [
            'PRD001', 'テスト商品', '2025-01-01', '2025-12-31',
            '文房具', '100', 'テスト説明文',
        ]
        for i, val in enumerate(expected_row):
            self.assertEqual(response.status_code, 200)
            self.assertEqual(rows[1][i], val)

    def test_export_excel_content_exact(self):
        """Excelエクスポートの中身を確認（ヘッダ + 1行目）"""
        url = reverse('product_mst:export_excel')
        response = self.client.get(url)

        # ステータス・Content-Type 確認
        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats', response['Content-Type'])

        # Excelファイルを読み込み
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # 1行目（ヘッダ）の確認
        expected_header = [
            'product_code', 'product_name', 'start_date', 'end_date',
            'product_category', 'price', 'description',
            'create_user', 'created_at', 'update_user', 'updated_at'
        ]
        actual_header = [cell.value for cell in ws[1]]
        self.assertEqual(actual_header, expected_header)

        # 2行目（データ）の確認
        expected_row = [
            'PRD001', 'テスト商品', '2025-01-01', '2025-12-31',
            '文房具', 100, 'テスト説明文',
            'tester', '2025-09-24', 'tester', '2025-09-24'
        ]
        actual_row = [cell.value for cell in ws[2]]
        # 日付は datetime 型になるので文字列に揃える
        actual_row = [v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else v for v in actual_row]

        self.assertEqual(actual_row, expected_row)

    # -------------------
    # Import
    # -------------------

    def test_import_csv_success_count_check(self):
        """CSVインポート成功 + 件数確認 + 登録内容確認"""
        url = reverse('product_mst:import_csv')
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        writer.writerow([
            'product_code', 'product_name', 'start_date', 'end_date',
            'product_category', 'price', 'description',
            'create_user', 'created_at', 'update_user', 'updated_at'
        ])
        writer.writerow([
            'NEW001', '新規商品', '2025-01-01', '2025-12-31',
            '文房具', '1000', '説明',
            'tester', '2025-01-01', 'tester', '2025-01-01'
        ])

        csv_file = SimpleUploadedFile(
            'test.csv',
            csv_content.getvalue().encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(url, {'file': csv_file}, format='multipart')

        # レスポンス確認
        self.assertEqual(response.status_code, 200)

        # 件数確認（初期1件 + 新規1件 = 2）
        self.assertEqual(Product.objects.count(), 2)

        # 登録内容を確認
        product = Product.objects.get(product_code='NEW001')
        self.assertEqual(product.product_name, '新規商品')
        self.assertEqual(product.start_date.strftime('%Y-%m-%d'), '2025-01-01')
        self.assertEqual(product.end_date.strftime('%Y-%m-%d'), '2025-12-31')
        self.assertEqual(product.product_category.product_category_name, '文房具')
        self.assertEqual(product.price, 1000)
        self.assertEqual(product.description, '説明')
        self.assertEqual(product.create_user.username, 'tester')
        self.assertEqual(product.update_user.username, 'tester')

    def test_import_csv_invalid_date_no_register(self):
        '''CSVインポート日付エラー → 件数変わらない'''
        url = reverse('product_mst:import_csv')
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)

        # ヘッダを expected_headers と完全一致させる
        writer.writerow([
            'product_code', 'product_name', 'start_date', 'end_date',
            'product_category', 'price', 'description',
            'create_user', 'created_at', 'update_user', 'updated_at'
        ])

        # データ行（終了日 < 開始日 → バリデーションエラーになる想定）
        writer.writerow([
            'ERR001', '不正商品', '2025-01-10', '2025-01-05',
            '文房具', '100', '説明',
            'tester', '2025-01-01', 'tester', '2025-01-01'
        ])

        csv_file = SimpleUploadedFile(
            'test.csv',
            csv_content.getvalue().encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(url, {'file': csv_file}, format='multipart')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(Product.objects.count(), 1)  # 件数変わらない
        self.assertFalse(Product.objects.filter(product_code='ERR001').exists())

    def test_import_csv_invalid_category(self):
        '''CSVインポート 異常系（存在しないカテゴリ）'''
        url = reverse('product_mst:import_csv')
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        writer.writerow(['product_code', 'product_name', 'start_date', 'end_date', 'product_category', 'price', 'description'])
        writer.writerow(['NEW200', '商品X', '2025-01-01', '2025-12-31', '存在しないカテゴリ', '100', '説明'])

        response = self.client.post(url, {'file': io.BytesIO(csv_content.getvalue().encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Product.objects.filter(product_code='NEW200').exists())

    def test_import_csv_invalid_price(self):
        '''CSVインポート 異常系（price が数値でない）'''
        url = reverse('product_mst:import_csv')
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        writer.writerow(['product_code', 'product_name', 'start_date', 'end_date', 'product_category', 'price', 'description'])
        writer.writerow(['NEW201', '商品Y', '2025-01-01', '2025-12-31', '文房具', 'abc', '説明'])

        response = self.client.post(url, {'file': io.BytesIO(csv_content.getvalue().encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Product.objects.filter(product_code='NEW201').exists())

    def test_import_csv_invalid_date_format(self):
        '''CSVインポート 異常系（日付フォーマット不正）'''
        url = reverse('product_mst:import_csv')
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        writer.writerow(['product_code', 'product_name', 'start_date', 'end_date', 'product_category', 'price', 'description'])
        writer.writerow(['NEW202', '商品Z', '2025/01/01', '2025-12-31', '文房具', '100', '説明'])

        response = self.client.post(url, {'file': io.BytesIO(csv_content.getvalue().encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Product.objects.filter(product_code='NEW202').exists())
